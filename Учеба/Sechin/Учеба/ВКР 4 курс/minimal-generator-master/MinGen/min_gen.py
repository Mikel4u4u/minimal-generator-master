from openpyxl.styles import Color, PatternFill, Font, Border
from MinGen.dataclasses import PromRow
import openpyxl


class MinGen:
    def __init__(self, signs_count, objects_count, data):
        self.signs_count = signs_count
        self.objects_count = objects_count
        self.data = self._form_data(data)
        self.prom_tab = self._make_prom_tab()
        self.result = []
        self.resulti = []
        self.gener = []
    @staticmethod
    def _form_data(data):
        formed_data = []
        for i in data:
            formed_row = []
            for j in i:
                formed_row.append(j)

            formed_data.append(formed_row)
        return formed_data

    def print_data(self):
        print(' '.join(self.chars))
        for i in range(self.objects_count):
            print(' '.join([str(j) for j in self.data[i]]))

    def print_result(self):
        for res in self.result:
            print(res)

    def print_prom_tab(self):
        data_to_print = [['X', 'K', 'X\'', 'X"' ]]
        for i in self.prom_tab:
            data_to_print.append([i.X_name, '+' if i.Key else '-', ''.join([str(j) for j in i.X_1]), i.X_2])
        for row in data_to_print:
            text = ''
            for i in row:
                text += i + ' ' * (10 - len(i))
            print(text)

    def _intersect(self, lst):
        if not lst:
            return [1] * self.signs_count
        res = []
        for i in range(len(lst[0])):
            if sum([lst[j][i] for j in range(len(lst))]) == len(lst):
                res.append(1)
            else:
                res.append(0)
        return res

    def _union(self, lst):
        if not lst:
            return [0] * self.signs_count
        res = []
        for i in range(len(lst[0])):
            if sum([lst[j][i] for j in range(len(lst))]) > 0:
                res.append(1)
            else:
                res.append(0)
        return res

    def _list_from_chars(self, chars):
        res = []
        for i in range(self.signs_count):
            if self.chars[i] in chars:
                res.append(1)
            else:
                res.append(0)
        return res

    def _chars_from_str(self, string):
        res = []
        for i in self.chars:
            if i in string:
                res.append(i)
        return res

    def _chars_from_list(self, lst):
        res = ""
        for i in range(len(lst)):
            if lst[i]:
                res += self.chars[i]
        return res if res else 'ø'

    def _get_objects_by_signs(self, signs_indexes):
        objects = []
        for j in range(len(self.data)):
            if sum([self.data[j][i] for i in signs_indexes]) == len(signs_indexes):
                objects.append(j)
        return objects

    def _get_objects_by_char(self, char):
        index = self.chars.index(char)
        return self._get_objects_by_signs([index])

    def _get_objects_by_x(self, x):
        signs_indexes = []
        for i in range(len(self.chars)):
            if x.find(self.chars[i]) != -1:
                signs_indexes.append(i)
        return self._get_objects_by_signs(signs_indexes)

    def _get_x_2_lst_from_objects(self, objects):
        obj_lists = [self.data[i] for i in range(self.objects_count) if i in objects]
        x_2_lst = self._intersect(obj_lists)
        return x_2_lst

    @property
    def chars(self):
        chars = []
        for i in range(self.signs_count):
            char = chr(ord('A') + i )
            chars.append(char)
        return chars

    def _make_prom_tab(self):
        prom_tab = []
        for char in self.chars:
            lst = self._list_from_chars(char)
            x_1 = self._get_objects_by_char(char)
            x_2_lst = self._get_x_2_lst_from_objects(x_1)
            x_2 = self._chars_from_list(x_2_lst)
            prom_tab.append(PromRow(char, x_2, x_1, x_2, x_2_lst,))
        return prom_tab

    def _approx(self, x: str):
        signs_x_2 = [prom_row.X_2_lst for prom_row in self.prom_tab if x.find(prom_row.X_name) != -1]
        return self._chars_from_list(self._union(signs_x_2))



    @staticmethod
    def _get_sublist(lst):
        return [[j for j in lst if j != i] for i in lst]

    def _is_equal_names(self, name1, name2):
        return set(self._chars_from_str(name1)) == set(self._chars_from_str(name2))

    def _name1_contain_name2(self, name1, name2):
        for char in self._chars_from_str(name2):
            if char not in name1:
                return False
        return True

    def _name1_del_name2(self, name1,name2):
        x_chars = self._chars_from_str(name2)
        for i in x_chars:
            name1 = name1.replace(i, '')
        return name1

    def _get_row_by_name(self, row_name):
        all_known_rows = self.prom_tab + self.result
        for row in all_known_rows:
            if self._is_equal_names(row.X_name, row_name):
                return row

    def _is_key(self, x: str,type):
        x_chars = self._chars_from_str(x)
        if len(x_chars) == 1:
            return True
        chars_subsets = self._get_sublist(x_chars)
        for chars_subset in chars_subsets:
            substring = ''.join(chars_subset)
            if not self._is_key(substring,type):
                return False
            substring_row = self._get_row_by_name(substring)
            if type == 'Apr':
                if self._name1_contain_name2(substring_row.X_Apr, x):
                    return False
            else:
                if self._name1_contain_name2(substring_row.X_2, x):
                    return False
        return True

    def check(self):
        for i in self.result:
            x_1 = self._get_objects_by_x(i.X_name)
            x_2_lst = self._get_x_2_lst_from_objects(x_1)
            x_2 = self._chars_from_list(x_2_lst)
            i.X_1 = x_1
            i.X_2_lst = x_2_lst
            i.X_2 = x_2
            i.Key = self._is_key(i.X_name,'X_2')

    def gen_next(self,iter):

        if iter == 1:
            t = self.prom_tab
        else:
            t = self.resulti

        for i in t:
            if i.Key:
                self.gener.append(i.X_name)

        y = Apgen(self.gener, len(self.gener))

        self.resulti.clear()
        self.gener.clear()

        for x in y:
            self.resulti.append(
                PromRow(
                    x,
                    self._approx(x),
                    0,
                    0,
                    0,
                    self._is_key(x,'Apr')
                )
            )
        self.result.extend(self.resulti)

    def gen_all(self):
        for iter in range(1,100):
            self.gen_next(iter)

    def save_all_to_excel(self):
        wb = openpyxl.Workbook()
        ws1 = wb.get_active_sheet()
        ws1.title = 'Input'
        ws1.append(['G\\M'] + self.chars)
        for i in range(self.objects_count):
            ws1.append([i] + self.data[i])
        ws2 = wb.create_sheet('Output')
        ws2.append([ 'X', 'K', 'X+', 'X\'', 'X"','З','Л.С.'])
        for i in self.prom_tab:
            ws2.append([i.X_name, '+' if i.Key else '-', i.X_Apr, ''.join(map(str, i.X_1)) if i.X_1 else 'ø',
                        i.X_2 if i.X_2 else 'ø','+' if i.X_name == i.X_2 else '-',
                        '-'])
        for i in self.result:
            ws2.append([ i.X_name, '+' if i.Key else '-',i.X_Apr, ''.join(map(str, i.X_1)) if i.X_1 else 'ø',
                         i.X_2 if i.X_2 else 'ø','+' if i.X_name == i.X_2 else '-',
                         '+' if i.X_Apr == i.X_2 else '-'])
        ws3 = wb.create_sheet('Minmax AR')
        ws3.append(['X -> ','X"'," ",'X -> ', 'X"\X ' ])

        for i in self.prom_tab:
            if i.Key and i.X_1  :
                ws3.append([i.X_name + "  ->",i.X_2  if i.X_2 else 'ø'," ",
                             i.X_name + "  ->" if self._name1_del_name2(i.X_2,i.X_name)
                             else '' ,self._name1_del_name2(i.X_2,i.X_name)])
        for i in self.result:
            if i.Key and i.X_1 and i.X_Apr != i.X_2 :
                ws3.append([i.X_name + "  ->", i.X_2 if i.X_2 else 'ø', " ",
                            i.X_name + "  ->" if self._name1_del_name2(i.X_2, i.X_name)
                            else '', self._name1_del_name2(i.X_2, i.X_name)])

        wb.save('output.xlsx')


def Apgen(Itemset, lenght):
    canditate = []

    for i in range(0, lenght):
        element = str(Itemset[i])
        for j in range(i + 1, lenght):
            element1 = str(Itemset[j])
            if element[0:(len(element) - 1)] == element1[0:(len(element1) - 1)]:
                unionset = element[0:(len(element) - 1)] + element1[len(element1) - 1] + element[
                    len(element) - 1]
                unionset = ''.join(sorted(unionset))
                canditate.append(unionset)
    return canditate
